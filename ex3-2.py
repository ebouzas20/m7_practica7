

from openpyxl import load_workbook

# ruta de nuestro archivo
filesheet = "./ex3-2.xlsx"

# creamos ell obejeto load_workbook
wb = load_workbook(filesheet)

datos = [('id', 'nombre', 'edad'),
   (0, "Jose", 35),
   (1, "Carlos", 27),
   (2, "Sofia", 24)]

sheet = wb.active

# Ingresamos el valor 56 en la celda 'A1'
sheet['A1'] = 56

# Ingresamos el valor 1845 en la celda 'B3'
sheet['B3'] = 1845


# recorremos las columnas y escribimos los datos
for row in datos:
 sheet.append(row)

# guardamos los cambios
wb.save(filesheet)
# seleccionamos el archivo
sheet = wb.active

# Obtenemos el valor de la celda A1
A1 = sheet['A1'].value

# Obtenemos el valor de la celda B5
B5 = sheet['B5'].value

# Obtenemos el valor de la celda C5
C5 = sheet['C5'].value

# Mostramos los valores 
celdas = [A1, B5, C5]
for valor in celdas:
 print(valor)